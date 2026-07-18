"""
Main pipeline: cleans a customer reviews CSV, runs AI sentiment analysis on
every review, and produces a client-ready Excel report with an embedded
sentiment chart.

Usage:
    python3 analyzer.py [path/to/reviews.csv]

If no path is given, defaults to sample_reviews.csv (see generate_sample_data.py).
"""

import os
import sys
from collections import Counter

import matplotlib
matplotlib.use("Agg")  # headless - we only need to save PNGs, not display them
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import ai_client

CHART_FILE = "sentiment_breakdown.png"
REPORT_FILE = "feedback_report.xlsx"

SENTIMENT_ORDER = ["positive", "negative", "neutral"]
SENTIMENT_COLORS = {"positive": "#2e7d32", "negative": "#c62828", "neutral": "#757575"}


def load_and_clean(csv_path):
    """Loads the CSV and drops duplicates, blank reviews, and stray whitespace."""
    df = pd.read_csv(csv_path)

    for col in ("customer_name", "review_text"):
        df[col] = df[col].fillna("").astype(str).str.strip()

    df = df.drop_duplicates()
    df = df[df["review_text"] != ""]
    df = df.reset_index(drop=True)

    return df


def run_analysis(df):
    """Runs sentiment/theme analysis on every review and adds the results as columns."""
    sentiments = []
    themes = []

    total = len(df)
    for i, review_text in enumerate(df["review_text"], start=1):
        sentiment, theme = ai_client.analyze_review(review_text)
        sentiments.append(sentiment)
        themes.append(theme)
        print(f"  Analyzed {i}/{total} reviews", end="\r")
    print()  # move past the progress line

    df["sentiment"] = sentiments
    df["theme"] = themes
    return df


def make_chart(sentiment_counts, output_path):
    """Builds a colored bar chart of sentiment counts and saves it as a PNG."""
    labels = [s for s in SENTIMENT_ORDER if sentiment_counts.get(s, 0) > 0] or SENTIMENT_ORDER
    counts = [sentiment_counts.get(s, 0) for s in labels]
    colors = [SENTIMENT_COLORS[s] for s in labels]

    fig, ax = plt.subplots(figsize=(6, 4))
    bars = ax.bar([s.capitalize() for s in labels], counts, color=colors)
    ax.set_title("Customer Sentiment Breakdown")
    ax.set_ylabel("Number of Reviews")
    ax.bar_label(bars, padding=3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def style_header_row(worksheet, row_num, num_columns):
    """Applies a dark background + bold white text to a header row."""
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, num_columns + 1):
        cell = worksheet.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def build_report(df, chart_path, output_path):
    """Assembles the two-sheet Excel workbook: Summary and All Reviews."""
    sentiment_counts = Counter(df["sentiment"])
    top_themes = Counter(df["theme"]).most_common(5)

    wb = Workbook()

    # --- Summary sheet ---
    summary = wb.active
    summary.title = "Summary"

    summary["A1"] = "AI-Powered Customer Feedback Report"
    summary["A1"].font = Font(size=16, bold=True)
    summary.merge_cells("A1:C1")

    summary["A3"] = "Total Reviews Analyzed"
    summary["A3"].font = Font(bold=True)
    summary["B3"] = len(df)

    row = 5
    summary.cell(row=row, column=1, value="Sentiment Breakdown").font = Font(bold=True)
    row += 1
    for sentiment in SENTIMENT_ORDER:
        summary.cell(row=row, column=1, value=sentiment.capitalize())
        summary.cell(row=row, column=2, value=sentiment_counts.get(sentiment, 0))
        row += 1

    row += 1
    summary.cell(row=row, column=1, value="Top 5 Themes").font = Font(bold=True)
    row += 1
    for theme, count in top_themes:
        summary.cell(row=row, column=1, value=theme)
        summary.cell(row=row, column=2, value=count)
        row += 1

    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 14
    summary.column_dimensions["C"].width = 14

    chart_image = XLImage(chart_path)
    chart_image.width, chart_image.height = 480, 320
    summary.add_image(chart_image, "D3")

    # --- All Reviews sheet ---
    reviews_sheet = wb.create_sheet("All Reviews")
    columns = list(df.columns)
    reviews_sheet.append(columns)
    style_header_row(reviews_sheet, 1, len(columns))

    for _, record in df.iterrows():
        reviews_sheet.append([record[col] for col in columns])

    column_widths = {
        "review_id": 10, "customer_name": 22, "review_text": 60,
        "sentiment": 12, "theme": 26,
    }
    for idx, col in enumerate(columns, start=1):
        reviews_sheet.column_dimensions[get_column_letter(idx)].width = column_widths.get(col, 20)

    wb.save(output_path)


def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else "sample_reviews.csv"

    if not os.path.exists(csv_path):
        print(f"Error: '{csv_path}' not found.")
        sys.exit(1)

    if ai_client.api_key_configured():
        print("GEMINI_API_KEY found - using the live Gemini API for sentiment analysis.")
    else:
        print("No GEMINI_API_KEY set - running in mock mode (keyword-based analysis, zero setup).")

    print(f"Loading and cleaning '{csv_path}'...")
    df = load_and_clean(csv_path)
    print(f"{len(df)} usable reviews after cleaning.")

    if len(df) == 0:
        print("No usable reviews found - nothing to analyze.")
        sys.exit(1)

    print("Running AI sentiment analysis...")
    df = run_analysis(df)

    print("Generating sentiment chart...")
    make_chart(Counter(df["sentiment"]), CHART_FILE)

    print("Building Excel report...")
    build_report(df, CHART_FILE, REPORT_FILE)

    print(f"Done! Report saved to '{REPORT_FILE}', chart saved to '{CHART_FILE}'.")


if __name__ == "__main__":
    main()
